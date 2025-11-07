import openpyxl

def check_totallist_structure():
    """Check the structure of TOTALLIST sheet to see all available columns."""
    print("Loading workbook...")
    wb = openpyxl.load_workbook('e2.xlsx', data_only=True)
    
    if 'TOTALLIST' not in wb.sheetnames:
        print("Error: TOTALLIST sheet not found!")
        return
    
    ws_totallist = wb['TOTALLIST']
    
    print(f"\nTOTALLIST Sheet Structure:")
    print(f"Total rows: {ws_totallist.max_row}")
    print(f"Total columns: {ws_totallist.max_column}")
    
    # Read header row
    print("\nHeader Row (Row 1):")
    headers = []
    for col in range(1, ws_totallist.max_column + 1):
        cell_value = ws_totallist.cell(1, col).value
        headers.append(cell_value)
        print(f"  Column {col} ({openpyxl.utils.get_column_letter(col)}): {cell_value}")
    
    # Read first few data rows to understand structure
    print("\nSample Data Rows (first 3 rows after header):")
    for row_idx in range(2, min(5, ws_totallist.max_row + 1)):
        print(f"\nRow {row_idx}:")
        for col in range(1, min(ws_totallist.max_column + 1, 10)):  # Show first 10 columns
            cell_value = ws_totallist.cell(row_idx, col).value
            header = headers[col - 1] if col <= len(headers) else f"Col{col}"
            print(f"  {header}: {cell_value}")
    
    # Check for MDB rows specifically
    print("\n\nSample MDB Row:")
    for row_idx in range(2, ws_totallist.max_row + 1):
        board_name = ws_totallist.cell(row_idx, 3).value  # Column C
        if board_name and 'MDB1' in str(board_name).upper():
            print(f"Row {row_idx} - MDB1:")
            for col in range(1, ws_totallist.max_column + 1):
                cell_value = ws_totallist.cell(row_idx, col).value
                header = headers[col - 1] if col <= len(headers) else f"Col{col}"
                if cell_value:
                    print(f"  {header}: {cell_value}")
            break

if __name__ == '__main__':
    check_totallist_structure()



