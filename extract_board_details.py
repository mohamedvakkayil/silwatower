import openpyxl
import json
import sys
import re

def extract_board_details(board_name):
    """Extract detailed data from a specific board sheet."""
    try:
        wb = openpyxl.load_workbook('e2.xlsx', data_only=True)
        
        if board_name not in wb.sheetnames:
            return {'error': f'Sheet "{board_name}" not found'}
        
        ws = wb[board_name]
        
        # Extract all data from the sheet
        details = {
            'board_name': board_name,
            'items': [],
            'summary': {}
        }
        
        # Find header row (usually row 1 or 2)
        header_row = None
        for row_idx in range(1, min(10, ws.max_row + 1)):
            row = ws[row_idx]
            # Check if this looks like a header row (contains common headers)
            if any(cell and isinstance(cell.value, str) and 
                   any(header in str(cell.value).upper() for header in ['ITEM', 'DESCRIPTION', 'QTY', 'UNIT', 'RATE', 'AMOUNT'])
                   for cell in row[:10]):
                header_row = row_idx
                break
        
        if header_row is None:
            header_row = 1
        
        # Extract headers
        headers = []
        for col in range(1, min(ws.max_column + 1, 20)):
            cell_value = ws.cell(header_row, col).value
            if cell_value:
                headers.append(str(cell_value).strip())
            else:
                headers.append(f'Column{col}')
        
        # Extract data rows
        for row_idx in range(header_row + 1, ws.max_row + 1):
            row_data = {}
            has_data = False
            
            for col_idx, header in enumerate(headers):
                if col_idx < ws.max_column:
                    cell_value = ws.cell(row_idx, col_idx + 1).value
                    if cell_value is not None:
                        row_data[header] = cell_value
                        has_data = True
            
            if has_data:
                details['items'].append(row_data)
        
        # Find summary values (NET TOTAL, NO OF UNITS, etc.)
        # Look for NET TOTAL and NO OF UNITS in the last 50 rows
        for row_idx in range(max(1, ws.max_row - 50), ws.max_row + 1):
            row = ws[row_idx]
            for col_idx in range(min(6, ws.max_column)):
                cell_value = row[col_idx].value
                if cell_value and isinstance(cell_value, str):
                    cell_upper = str(cell_value).upper().strip()
                    # Look for NET TOTAL (must be exact match or contain NET TOTAL)
                    if 'NET TOTAL' in cell_upper and 'NET TOTAL' not in details['summary']:
                        # Try to get the value from amount column (usually column F, index 5)
                        # Also check adjacent columns in case structure varies
                        for check_col in [6, 5, 7, 4]:
                            if ws.max_column >= check_col:
                                amount_value = ws.cell(row_idx, check_col).value
                                if amount_value is not None:
                                    try:
                                        # Try to extract number from string if needed
                                        if isinstance(amount_value, str):
                                            # Extract numeric part (remove text like "kVAR", "kW", etc.)
                                            num_match = re.search(r'[\d,]+\.?\d*', amount_value.replace(',', ''))
                                            if num_match:
                                                amount_value = float(num_match.group())
                                            else:
                                                continue
                                        else:
                                            amount_value = float(amount_value)
                                        details['summary']['net_total'] = amount_value
                                        break
                                    except (ValueError, TypeError):
                                        continue
                    # Look for NO OF UNITS (must be exact match)
                    elif ('NO OF UNITS' in cell_upper or 'NO OF ITEMS' in cell_upper) and 'no_of_units' not in details['summary']:
                        # Try to get the value from amount column (usually column F, index 5)
                        # Also check adjacent columns in case structure varies
                        for check_col in [6, 5, 7, 4]:
                            if ws.max_column >= check_col:
                                units_value = ws.cell(row_idx, check_col).value
                                if units_value is not None:
                                    try:
                                        # Try to extract number from string if needed
                                        if isinstance(units_value, str):
                                            # Extract numeric part (remove text)
                                            num_match = re.search(r'[\d,]+\.?\d*', units_value.replace(',', ''))
                                            if num_match:
                                                units_value = float(num_match.group())
                                            else:
                                                continue
                                        else:
                                            units_value = float(units_value)
                                        details['summary']['no_of_units'] = units_value
                                        break
                                    except (ValueError, TypeError):
                                        continue
        
        return details
        
    except Exception as e:
        return {'error': str(e)}

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(json.dumps({'error': 'Board name required'}))
        sys.exit(1)
    
    board_name = sys.argv[1]
    result = extract_board_details(board_name)
    print(json.dumps(result, indent=2, default=str))

