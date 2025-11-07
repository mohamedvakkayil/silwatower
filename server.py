from flask import Flask, send_from_directory, jsonify
from flask_cors import CORS
import openpyxl
import os
import re
from datetime import datetime
from collections import OrderedDict

app = Flask(__name__, static_folder='.')
CORS(app)  # Enable CORS for all routes

def extract_dashboard_data():
    """Extract dashboard data directly from Excel file."""
    try:
        wb = openpyxl.load_workbook('e2.xlsx', data_only=True)
        
        if 'TOTALLIST' not in wb.sheetnames:
            return {'error': 'TOTALLIST sheet not found'}
        
        ws_totallist = wb['TOTALLIST']
        
        # Column indices (1-based) - Based on header: NumTag, KIND, MDB, SMDB, Itemdrop, Load, NO OF ITEMS, Estimate
        board_name_col = 5  # Column E (Itemdrop - board name)
        estimate_col = 8    # Column H (Estimate)
        load_col = 6        # Column F (Load)
        items_col = 7       # Column G (NO OF ITEMS)
        kind_col = 2        # Column B (KIND)
        mdb_col = 3         # Column C (MDB)
        
        mdb_boards = []
        all_boards = []
        
        # Process rows starting from row 2
        for row_idx in range(2, ws_totallist.max_row + 1):
            board_name_cell = ws_totallist.cell(row_idx, board_name_col)
            estimate_cell = ws_totallist.cell(row_idx, estimate_col)
            load_cell = ws_totallist.cell(row_idx, load_col) if load_col <= ws_totallist.max_column else None
            items_cell = ws_totallist.cell(row_idx, items_col) if items_col <= ws_totallist.max_column else None
            kind_cell = ws_totallist.cell(row_idx, kind_col) if kind_col <= ws_totallist.max_column else None
            mdb_cell = ws_totallist.cell(row_idx, mdb_col) if mdb_col <= ws_totallist.max_column else None
            
            board_name = board_name_cell.value
            estimate = estimate_cell.value
            load = load_cell.value if load_cell else None
            items = items_cell.value if items_cell else None
            kind = kind_cell.value if kind_cell else None
            mdb = mdb_cell.value if mdb_cell else None
            
            if not board_name or str(board_name).strip() == '':
                continue
            
            board_name = str(board_name).strip()
            board_name_upper = board_name.upper()
            
            # Check if it's one of the 4 main MDBs
            # Also check by KIND column - if KIND is 'MDB', it's a main MDB
            is_main_mdb = (
                kind and str(kind).upper() == 'MDB' and (
                    board_name_upper == 'MDB1' or 
                    board_name_upper == 'MDB2' or 
                    board_name_upper == 'MDB3' or 
                    board_name_upper == 'MDB4' or
                    board_name_upper == 'MDB.GF.04' or
                    board_name_upper.startswith('MDB1 ') or
                    board_name_upper.startswith('MDB2 ') or
                    board_name_upper.startswith('MDB3 ') or
                    board_name_upper.startswith('MDB4 ')
                )
            )
            
            try:
                estimate_value = float(estimate) if estimate is not None else 0
                
                # Parse load value (might have "kW" suffix or be a string)
                load_value = 0
                if load is not None:
                    if isinstance(load, (int, float)):
                        load_value = float(load)
                    elif isinstance(load, str):
                        # Remove "kW" and whitespace, then parse
                        load_str = load.replace('kW', '').replace('kW', '').strip()
                        try:
                            load_value = float(load_str)
                        except ValueError:
                            load_value = 0
                    else:
                        load_value = 0
                
                items_value = int(items) if items is not None else 0
                
                board_data = {
                    'name': board_name,
                    'estimate': estimate_value,
                    'load': load_value,
                    'items': items_value,
                    'kind': str(kind) if kind else None,
                    'mdb': str(mdb) if mdb else None
                }
                
                all_boards.append(board_data)
                
                if is_main_mdb:
                    mdb_boards.append(board_data)
            except (ValueError, TypeError) as e:
                continue
        
        # Sort MDB boards
        def sort_key(x):
            name = x['name'].upper()
            if name.startswith('MDB1'):
                return (0, name)
            elif name.startswith('MDB2'):
                return (1, name)
            elif name.startswith('MDB3'):
                return (2, name)
            elif 'MDB.GF.04' in name:
                return (3, name)
            else:
                return (4, name)
        
        mdb_boards.sort(key=sort_key)
        
        # Calculate totals
        mdb_total_estimate = sum(item['estimate'] for item in mdb_boards)
        mdb_total_load = sum(item['load'] for item in mdb_boards)
        mdb_total_items = sum(item['items'] for item in mdb_boards)
        
        all_total_estimate = sum(item['estimate'] for item in all_boards)
        all_total_load = sum(item['load'] for item in all_boards)
        all_total_items = sum(item['items'] for item in all_boards)
        
        return {
            'mdbBoards': mdb_boards,
            'totalEstimate': mdb_total_estimate,
            'totalLoad': mdb_total_load,
            'totalItems': mdb_total_items,
            'mdbCount': len(mdb_boards),
            'allBoards': all_boards,
            'allBoardsTotal': all_total_estimate,
            'allBoardsTotalLoad': all_total_load,
            'allBoardsTotalItems': all_total_items,
            'allBoardsCount': len(all_boards),
            'lastUpdated': datetime.now().isoformat()
        }
    except Exception as e:
        return {'error': str(e)}

@app.route('/')
def index():
    """Serve the dashboard HTML."""
    return send_from_directory('.', 'dashboard.html')

@app.route('/api/dashboard-data')
def dashboard_data():
    """API endpoint to get dashboard data from Excel."""
    data = extract_dashboard_data()
    return jsonify(data)

@app.route('/api/board-details')
def board_details():
    """API endpoint to get board details from Excel."""
    from flask import request
    
    board_name = request.args.get('name')
    if not board_name:
        return jsonify({'error': 'Board name is required'}), 400
    
    try:
        wb = openpyxl.load_workbook('e2.xlsx', data_only=True)
        
        if board_name not in wb.sheetnames:
            return jsonify({'error': f'Board sheet "{board_name}" not found'}), 404
        
        # Get board metadata from TOTALLIST sheet
        board_metadata = {}
        if 'TOTALLIST' in wb.sheetnames:
            ws_totallist = wb['TOTALLIST']
            # Column indices: NumTag=1, KIND=2, MDB=3, SMDB=4, Itemdrop=5, Load=6, NO OF ITEMS=7, Estimate=8
            for row_idx in range(2, ws_totallist.max_row + 1):
                itemdrop = ws_totallist.cell(row_idx, 5).value  # Column E (Itemdrop - board name)
                if itemdrop and str(itemdrop).strip() == board_name:
                    kind = ws_totallist.cell(row_idx, 2).value  # Column B (KIND)
                    mdb = ws_totallist.cell(row_idx, 3).value   # Column C (MDB)
                    smdb = ws_totallist.cell(row_idx, 4).value  # Column D (SMDB)
                    load = ws_totallist.cell(row_idx, 6).value # Column F (Load)
                    
                    # Parse load value safely (handle strings with units like "300.00 kVAR" or "100.5 kW")
                    load_value = None
                    if load is not None:
                        try:
                            if isinstance(load, str):
                                # Extract numeric part from string (remove units like "kW", "kVAR", etc.)
                                num_match = re.search(r'[\d,]+\.?\d*', load.replace(',', ''))
                                if num_match:
                                    load_value = float(num_match.group())
                            else:
                                load_value = float(load)
                        except (ValueError, TypeError):
                            load_value = None
                    
                    board_metadata = {
                        'kind': str(kind).strip() if kind else None,
                        'mdb': str(mdb).strip() if mdb else None,
                        'smdb': str(smdb).strip() if smdb else None,
                        'load': load_value
                    }
                    break
        
        ws = wb[board_name]
        
        # Find header row (look for row with common headers like ITEM, QTY, PRICE, AMOUNT)
        header_row = 1
        for row_idx in range(1, min(10, ws.max_row + 1)):
            row_values = []
            for col_idx in range(1, min(ws.max_column + 1, 15)):
                cell_value = ws.cell(row_idx, col_idx).value
                row_values.append(str(cell_value).upper() if cell_value else '')
            
            # Check if this looks like a header row
            has_header_keywords = any(keyword in ' '.join(row_values) for keyword in 
                                     ['ITEM', 'QTY', 'PRICE', 'AMOUNT', 'BRAND', 'DESCRIPTION'])
            
            if has_header_keywords:
                header_row = row_idx
                break
        
        # Define preferred column order (most common columns first)
        # Note: 'back' is excluded as it's just a hyperlink, not a data column
        preferred_order = ['BRAND', 'ITEM', 'DESCRIPTION', 'PRICE', 'QTY', 'QUANTITY', 'AMOUNT']
        
        # Columns to exclude (not data columns)
        excluded_columns = ['back', 'BACK', 'list', 'LIST']
        
        # Extract headers - map each column to its header
        # Exclude 'back' and 'list' columns as they're just hyperlinks/formatting
        headers = {}
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(header_row, col).value
            if cell_value:
                header_name = str(cell_value).strip()
                # Skip excluded columns (hyperlinks, not data)
                if header_name and header_name.upper() not in [ex.upper() for ex in excluded_columns]:
                    headers[col] = header_name
        
        # If no headers found, use default column names
        if not headers:
            for col in range(1, min(ws.max_column + 1, 20)):
                headers[col] = f'Column{col}'
        
        # Get all headers and determine their order
        all_headers_list = list(headers.values())
        all_headers_dict_upper = {h.upper(): h for h in all_headers_list}
        
        # Create ordered header list: preferred first, then others alphabetically
        ordered_headers = []
        used_headers = set()
        
        # Add preferred headers in order (case-insensitive match)
        for pref_header in preferred_order:
            pref_upper = pref_header.upper()
            if pref_upper in all_headers_dict_upper and pref_upper not in used_headers:
                original_header = all_headers_dict_upper[pref_upper]
                ordered_headers.append(original_header)
                used_headers.add(pref_upper)
        
        # Add remaining headers alphabetically (by original case)
        remaining = [h for h_upper, h in all_headers_dict_upper.items() if h_upper not in used_headers]
        ordered_headers.extend(sorted(remaining, key=lambda x: x.upper()))
        
        # Extract data rows in the preferred order
        items = []
        for row_idx in range(header_row + 1, ws.max_row + 1):
            row_data = OrderedDict()
            has_data = False
            
            # Map columns in the preferred order
            for header_name in ordered_headers:
                # Find the column index for this header
                col_idx = None
                for col, h in headers.items():
                    if h == header_name:
                        col_idx = col
                        break
                
                if col_idx:
                    cell_value = ws.cell(row_idx, col_idx).value
                    row_data[header_name] = cell_value
                    if cell_value is not None and str(cell_value).strip() != '':
                        has_data = True
            
            # Only add rows that have some data
            if has_data:
                # Keep as OrderedDict to ensure order is preserved through JSON serialization
                items.append(row_data)
        
        # Find summary data (NET TOTAL, NO OF UNITS)
        summary = {}
        for row_idx in range(max(1, ws.max_row - 50), ws.max_row + 1):
            row = ws[row_idx]
            if len(row) > 2 and row[2].value:
                item_str = str(row[2].value).upper().strip()
                if 'NET TOTAL' in item_str and 'net_total' not in summary:
                    # Try multiple columns to find the value
                    for check_col in [5, 4, 6, 3]:
                        if len(row) > check_col:
                            value = row[check_col].value
                            if value is not None:
                                try:
                                    # Handle string values with units (e.g., "300.00 kVAR")
                                    if isinstance(value, str):
                                        num_match = re.search(r'[\d,]+\.?\d*', value.replace(',', ''))
                                        if num_match:
                                            value = float(num_match.group())
                                        else:
                                            continue
                                    else:
                                        value = float(value)
                                    summary['net_total'] = value
                                    break
                                except (ValueError, TypeError):
                                    continue
                elif ('NO OF UNITS' in item_str or 'NO OF ITEMS' in item_str) and 'no_of_units' not in summary:
                    # Try multiple columns to find the value
                    for check_col in [5, 4, 6, 3]:
                        if len(row) > check_col:
                            value = row[check_col].value
                            if value is not None:
                                try:
                                    # Handle string values with units
                                    if isinstance(value, str):
                                        num_match = re.search(r'[\d,]+\.?\d*', value.replace(',', ''))
                                        if num_match:
                                            value = float(num_match.group())
                                        else:
                                            continue
                                    else:
                                        value = float(value)
                                    summary['no_of_units'] = value
                                    break
                                except (ValueError, TypeError):
                                    continue
        
        # Convert OrderedDict items to regular dicts (Python 3.7+ preserves order)
        # But keep the order by ensuring we serialize in the right order
        items_dicts = []
        for item in items:
            if isinstance(item, OrderedDict):
                items_dicts.append(dict(item))
            else:
                items_dicts.append(item)
        
        return jsonify({
            'name': board_name,
            'metadata': board_metadata,
            'summary': summary,
            'items': items_dicts
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/<path:path>')
def serve_static(path):
    """Serve static files."""
    return send_from_directory('.', path)

if __name__ == '__main__':
    import sys
    
    # Try to get port from command line argument, or use default
    port = 5001  # Default to 5001 to avoid conflict with macOS AirPlay
    if len(sys.argv) > 1:
        try:
            port = int(sys.argv[1])
        except ValueError:
            print(f"Invalid port number: {sys.argv[1]}. Using default port 5001.")
    
    print("Starting dashboard server...")
    print(f"Dashboard will be available at: http://localhost:{port}")
    print(f"API endpoint: http://localhost:{port}/api/dashboard-data")
    print(f"\nNote: If port {port} is in use, you can specify a different port:")
    print(f"  python3 server.py 8000")
    
    try:
        app.run(debug=True, host='0.0.0.0', port=port)
    except OSError as e:
        if "Address already in use" in str(e):
            print(f"\nError: Port {port} is already in use.")
            print(f"Try running with a different port:")
            print(f"  python3 server.py 8000")
            print(f"\nOr on macOS, disable AirPlay Receiver:")
            print(f"  System Preferences -> General -> AirDrop & Handoff -> AirPlay Receiver")
        else:
            raise

