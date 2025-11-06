import openpyxl
from openpyxl.utils import get_column_letter

def get_board_total(board_sheet_name, wb):
    """Get the NET TOTAL value from a board sheet."""
    try:
        if board_sheet_name not in wb.sheetnames:
            return None
        
        ws = wb[board_sheet_name]
        
        # Find NET TOTAL row (usually near the bottom)
        for row_idx in range(ws.max_row - 20, ws.max_row + 1):
            row = ws[row_idx]
            # Check column C (ITEM column, index 2) for "NET TOTAL"
            if row[2].value and 'NET TOTAL' in str(row[2].value).upper():
                # Get value from column F (AMOUNT column, index 5)
                net_total = row[5].value
                if net_total is not None:
                    try:
                        return float(net_total)
                    except (ValueError, TypeError):
                        return None
        return None
    except Exception as e:
        return None

def get_no_of_units(board_sheet_name, wb):
    """Get the NO OF UNITS value from a board sheet (this is the value for NO OF ITEMS)."""
    try:
        if board_sheet_name not in wb.sheetnames:
            return None
        
        ws = wb[board_sheet_name]
        
        # Find NO OF UNITS row (usually row 156)
        for row_idx in range(ws.max_row - 20, ws.max_row + 1):
            row = ws[row_idx]
            # Check column C (ITEM column, index 2) for "NO OF UNITS"
            if row[2].value and 'NO OF UNITS' in str(row[2].value).upper():
                # Get value from column F (AMOUNT column, index 5)
                no_of_units = row[5].value
                if no_of_units is not None:
                    try:
                        return float(no_of_units)
                    except (ValueError, TypeError):
                        return None
        return None
    except Exception as e:
        return None

def update_estimates():
    """Update the Estimate and NO OF ITEMS columns in TOTALLIST sheet with values from each board sheet."""
    print("Loading workbook...")
    wb = openpyxl.load_workbook('e2.xlsx', data_only=True)
    
    if 'TOTALLIST' not in wb.sheetnames:
        print("Error: TOTALLIST sheet not found!")
        return
    
    ws_totallist = wb['TOTALLIST']
    
    # Column E is Itemdrop which contains board names (index 4, column 5)
    board_name_col = 5
    
    # Column G is NO OF ITEMS (index 6, column 7)
    items_col = 7
    
    # Column H is Estimate (index 7, column 8)
    estimate_col = 8
    
    print(f"Found {ws_totallist.max_row} rows in TOTALLIST sheet")
    print(f"Processing rows from 2 to {ws_totallist.max_row}...")
    
    updated_estimates = 0
    updated_items = 0
    not_found_count = 0
    
    # Process each row (starting from row 2, row 1 is header)
    for row_idx in range(2, ws_totallist.max_row + 1):
        board_name_cell = ws_totallist.cell(row_idx, board_name_col)
        board_name = board_name_cell.value
        
        if not board_name or str(board_name).strip() == '':
            continue
        
        board_name = str(board_name).strip()
        estimate_cell = ws_totallist.cell(row_idx, estimate_col)
        items_cell = ws_totallist.cell(row_idx, items_col)
        
        # Get the total from the board sheet
        net_total = get_board_total(board_name, wb)
        no_of_units = get_no_of_units(board_name, wb)
        
        if net_total is not None:
            estimate_cell.value = net_total
            updated_estimates += 1
        
        if no_of_units is not None:
            items_cell.value = no_of_units
            updated_items += 1
        
        if net_total is None and no_of_units is None:
            not_found_count += 1
            if not_found_count <= 5:  # Show first 5 warnings
                print(f"  Row {row_idx}: Could not find data for board '{board_name}'")
        
        if (updated_estimates + updated_items) % 20 == 0:
            print(f"  Processed {row_idx - 1} rows...")
    
    print(f"\nSummary:")
    print(f"  Successfully updated estimates: {updated_estimates} rows")
    print(f"  Successfully updated NO OF ITEMS: {updated_items} rows")
    print(f"  Not found/errors: {not_found_count} rows")
    
    # Save the workbook
    print("\nSaving workbook...")
    wb.save('e2.xlsx')
    print("Done! Estimates and NO OF ITEMS have been updated in the TOTALLIST sheet.")

if __name__ == '__main__':
    update_estimates()


