import openpyxl
import json
import os

def extract_mdb_data():
    """Extract MDB data from TOTALLIST sheet and convert to JSON."""
    print("Loading workbook...")
    wb = openpyxl.load_workbook('e2.xlsx', data_only=True)
    
    if 'TOTALLIST' not in wb.sheetnames:
        print("Error: TOTALLIST sheet not found!")
        return
    
    ws_totallist = wb['TOTALLIST']
    
    # Find MDB rows (likely contain "MDB" in the board name)
    mdb_data = []
    
    # Column indices (1-based)
    # Assuming structure: A=Serial, B=?, C=Itemdrop (Board Name), D=Estimate, etc.
    board_name_col = 3  # Column C
    estimate_col = 4    # Column D
    
    print(f"Scanning {ws_totallist.max_row} rows in TOTALLIST sheet...")
    
    # Process rows starting from row 2 (assuming row 1 is header)
    for row_idx in range(2, ws_totallist.max_row + 1):
        board_name_cell = ws_totallist.cell(row_idx, board_name_col)
        estimate_cell = ws_totallist.cell(row_idx, estimate_col)
        
        board_name = board_name_cell.value
        estimate = estimate_cell.value
        
        if not board_name or str(board_name).strip() == '':
            continue
        
        board_name = str(board_name).strip()
        
        # Check if it's one of the 4 main MDBs (MDB1, MDB2, MDB3, MDB.GF.04)
        # Pattern: MDB followed by number, or MDB.GF.04
        board_name_upper = board_name.upper()
        is_main_mdb = (
            board_name_upper == 'MDB1' or 
            board_name_upper == 'MDB2' or 
            board_name_upper == 'MDB3' or 
            board_name_upper == 'MDB.GF.04' or
            board_name_upper.startswith('MDB1 ') or
            board_name_upper.startswith('MDB2 ') or
            board_name_upper.startswith('MDB3 ')
        )
        
        if is_main_mdb:
            try:
                estimate_value = float(estimate) if estimate is not None else 0
                mdb_data.append({
                    'name': board_name,
                    'estimate': estimate_value
                })
                print(f"  Found Main MDB: {board_name} - {estimate_value:,.2f}")
            except (ValueError, TypeError):
                print(f"  Warning: Could not parse estimate for {board_name}")
    
    # Sort by name: MDB1, MDB2, MDB3, MDB.GF.04
    def sort_key(x):
        name = x['name'].upper()
        if name.startswith('MDB1'):
            return (0, name)
        elif name.startswith('MDB2'):
            return (1, name)
        elif name.startswith('MDB3'):
            return (2, name)
        elif 'MDB.GF.04' in name:
            return (3, name)
        else:
            return (4, name)
    
    mdb_data.sort(key=sort_key)
    
    # Calculate total
    total_estimate = sum(item['estimate'] for item in mdb_data)
    
    # Create output structure
    output = {
        'mdb_boards': mdb_data,
        'total_estimate': total_estimate,
        'count': len(mdb_data)
    }
    
    # Save to JSON
    output_file = 'mdb_data.json'
    with open(output_file, 'w') as f:
        json.dump(output, f, indent=2)
    
    print(f"\nSummary:")
    print(f"  Found {len(mdb_data)} MDB boards")
    print(f"  Total Estimate: {total_estimate:,.2f}")
    print(f"  Data saved to {output_file}")
    
    return output

if __name__ == '__main__':
    extract_mdb_data()

