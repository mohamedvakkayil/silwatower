import openpyxl
import json

def extract_all_boards():
    """Extract all boards from TOTALLIST sheet with all available columns."""
    print("Loading workbook...")
    wb = openpyxl.load_workbook('e2.xlsx', data_only=True)
    
    if 'TOTALLIST' not in wb.sheetnames:
        print("Error: TOTALLIST sheet not found!")
        return
    
    ws_totallist = wb['TOTALLIST']
    
    # Find all boards
    all_boards = []
    main_mdb_boards = []
    
    # Column indices (1-based) - Updated based on actual structure
    numtag_col = 1      # Column A
    kind_col = 2        # Column B
    mdb_col = 3         # Column C
    smdb_col = 4        # Column D
    itemdrop_col = 5    # Column E (Board Name)
    load_col = 6        # Column F
    items_col = 7       # Column G
    estimate_col = 8    # Column H
    
    print(f"Scanning {ws_totallist.max_row} rows in TOTALLIST sheet...")
    
    # Process rows starting from row 2 (assuming row 1 is header)
    for row_idx in range(2, ws_totallist.max_row + 1):
        board_name_cell = ws_totallist.cell(row_idx, itemdrop_col)
        board_name = board_name_cell.value
        
        if not board_name or str(board_name).strip() == '':
            continue
        
        board_name = str(board_name).strip()
        
        # Extract all fields
        numtag = ws_totallist.cell(row_idx, numtag_col).value
        kind = ws_totallist.cell(row_idx, kind_col).value
        mdb = ws_totallist.cell(row_idx, mdb_col).value
        smdb = ws_totallist.cell(row_idx, smdb_col).value
        load = ws_totallist.cell(row_idx, load_col).value
        items = ws_totallist.cell(row_idx, items_col).value
        estimate = ws_totallist.cell(row_idx, estimate_col).value
        
        try:
            estimate_value = float(estimate) if estimate is not None else 0
            items_value = int(items) if items is not None and str(items).isdigit() else (float(items) if items is not None else 0)
            
            # Parse load (remove "kW" if present)
            load_value = None
            if load:
                load_str = str(load).replace('kW', '').replace(' ', '').strip()
                try:
                    load_value = float(load_str)
                except (ValueError, TypeError):
                    load_value = None
            
            board_data = {
                'name': board_name,
                'numtag': numtag if numtag is not None else None,
                'kind': str(kind).strip() if kind else None,
                'mdb': str(mdb).strip() if mdb else None,
                'smdb': str(smdb).strip() if smdb else None,
                'load': load_value,
                'items': items_value,
                'estimate': estimate_value
            }
            
            all_boards.append(board_data)
            
            # Check if it's one of the 4 main MDBs
            # Main MDBs are identified by KIND='MDB' and the board name matches MDB1, MDB2, MDB3, or MDB.GF.04
            board_name_upper = board_name.upper()
            kind_upper = str(kind).upper() if kind else ''
            
            is_main_mdb = (
                kind_upper == 'MDB' and (
                    board_name_upper == 'MDB1' or 
                    board_name_upper == 'MDB2' or 
                    board_name_upper == 'MDB3' or 
                    board_name_upper == 'MDB4' or
                    board_name_upper == 'MDB.GF.04' or
                    'MDB.GF.04' in board_name_upper or
                    (mdb and str(mdb).upper() in ['MDB1', 'MDB2', 'MDB3', 'MDB.GF.04', 'MDB4'])
                )
            )
            
            if is_main_mdb:
                main_mdb_boards.append(board_data)
                
        except (ValueError, TypeError) as e:
            print(f"  Warning: Could not parse data for {board_name}: {e}")
    
    # Sort main MDBs
    def sort_key(x):
        name = x['name'].upper()
        if name.startswith('MDB1') or name == 'MDB1':
            return (0, name)
        elif name.startswith('MDB2') or name == 'MDB2':
            return (1, name)
        elif name.startswith('MDB3') or name == 'MDB3':
            return (2, name)
        elif name.startswith('MDB4') or name == 'MDB4' or 'MDB.GF.04' in name:
            return (3, name)
        else:
            return (4, name)
    
    main_mdb_boards.sort(key=sort_key)
    
    # Sort all boards by name
    all_boards.sort(key=lambda x: x['name'])
    
    # Calculate totals and statistics
    main_total = sum(item['estimate'] for item in main_mdb_boards)
    all_total = sum(item['estimate'] for item in all_boards)
    
    # Calculate total load for main MDBs
    main_total_load = sum(item['load'] for item in main_mdb_boards if item['load'] is not None)
    all_total_load = sum(item['load'] for item in all_boards if item['load'] is not None)
    
    # Count total items
    main_total_items = sum(item['items'] for item in main_mdb_boards)
    all_total_items = sum(item['items'] for item in all_boards)
    
    # Group boards by MDB
    boards_by_mdb = {}
    for board in all_boards:
        mdb_key = board['mdb'] or 'Unassigned'
        if mdb_key not in boards_by_mdb:
            boards_by_mdb[mdb_key] = []
        boards_by_mdb[mdb_key].append(board)
    
    # Create output structure for main MDBs
    mdb_output = {
        'mdb_boards': main_mdb_boards,
        'total_estimate': main_total,
        'total_load': main_total_load,
        'total_items': main_total_items,
        'count': len(main_mdb_boards)
    }
    
    # Create output structure for all boards
    all_boards_output = {
        'all_boards': all_boards,
        'total_estimate': all_total,
        'total_load': all_total_load,
        'total_items': all_total_items,
        'count': len(all_boards),
        'main_mdbs': main_mdb_boards,
        'main_mdb_total': main_total,
        'boards_by_mdb': boards_by_mdb
    }
    
    # Save to JSON files
    with open('mdb_data.json', 'w') as f:
        json.dump(mdb_output, f, indent=2)
    
    with open('all_boards_data.json', 'w') as f:
        json.dump(all_boards_output, f, indent=2)
    
    print(f"\nSummary:")
    print(f"  Main MDBs: {len(main_mdb_boards)}")
    print(f"  Main MDB Total: {main_total:,.2f} AED")
    print(f"  Main MDB Total Load: {main_total_load:,.2f} kW")
    print(f"  All Boards: {len(all_boards)}")
    print(f"  All Boards Total: {all_total:,.2f} AED")
    print(f"  All Boards Total Load: {all_total_load:,.2f} kW")
    print(f"  Total Items: {all_total_items}")
    print(f"  Data saved to mdb_data.json and all_boards_data.json")
    
    return mdb_output, all_boards_output

if __name__ == '__main__':
    extract_all_boards()
