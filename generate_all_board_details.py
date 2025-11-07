#!/usr/bin/env python3
"""
Generate JSON files for all board details.
This script extracts details for all boards and saves them as individual JSON files
so they can be served as static files for online deployment.
"""

import openpyxl
import json
import os
import re
from collections import OrderedDict

def extract_board_details(board_name, wb):
    """Extract detailed data from a specific board sheet."""
    try:
        if board_name not in wb.sheetnames:
            return {'error': f'Sheet "{board_name}" not found'}
        
        ws = wb[board_name]
        
        # Get board metadata from TOTALLIST sheet
        board_metadata = {}
        if 'TOTALLIST' in wb.sheetnames:
            ws_totallist = wb['TOTALLIST']
            for row_idx in range(2, ws_totallist.max_row + 1):
                itemdrop = ws_totallist.cell(row_idx, 5).value
                if itemdrop and str(itemdrop).strip() == board_name:
                    kind = ws_totallist.cell(row_idx, 2).value
                    mdb = ws_totallist.cell(row_idx, 3).value
                    smdb = ws_totallist.cell(row_idx, 4).value
                    load = ws_totallist.cell(row_idx, 6).value
                    
                    # Parse load value safely
                    load_value = None
                    if load is not None:
                        try:
                            if isinstance(load, str):
                                num_match = re.search(r'[\d,]+\.?\d*', load.replace(',', ''))
                                if num_match:
                                    load_value = float(num_match.group())
                            else:
                                load_value = float(load)
                        except (ValueError, TypeError):
                            load_value = None
                    
                    board_metadata = {
                        'kind': str(kind).strip() if kind else None,
                        'mdb': str(mdb).strip() if mdb else None,
                        'smdb': str(smdb).strip() if smdb else None,
                        'load': load_value
                    }
                    break
        
        # Find header row
        header_row = 1
        for row_idx in range(1, min(10, ws.max_row + 1)):
            row_values = []
            for col_idx in range(1, min(ws.max_column + 1, 15)):
                cell_value = ws.cell(row_idx, col_idx).value
                row_values.append(str(cell_value).upper() if cell_value else '')
            
            has_header_keywords = any(keyword in ' '.join(row_values) for keyword in 
                                     ['ITEM', 'QTY', 'PRICE', 'AMOUNT', 'BRAND', 'DESCRIPTION'])
            
            if has_header_keywords:
                header_row = row_idx
                break
        
        # Define preferred column order
        preferred_order = ['BRAND', 'ITEM', 'DESCRIPTION', 'PRICE', 'QTY', 'QUANTITY', 'AMOUNT']
        excluded_columns = ['back', 'BACK', 'list', 'LIST']
        
        # Extract headers
        headers = {}
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(header_row, col).value
            if cell_value:
                header_name = str(cell_value).strip()
                if header_name and header_name.upper() not in [ex.upper() for ex in excluded_columns]:
                    headers[col] = header_name
        
        if not headers:
            for col in range(1, min(ws.max_column + 1, 20)):
                headers[col] = f'Column{col}'
        
        # Get all headers and determine their order
        all_headers_list = list(headers.values())
        all_headers_dict_upper = {h.upper(): h for h in all_headers_list}
        
        # Create ordered header list
        ordered_headers = []
        used_headers = set()
        
        for pref_header in preferred_order:
            pref_upper = pref_header.upper()
            if pref_upper in all_headers_dict_upper and pref_upper not in used_headers:
                original_header = all_headers_dict_upper[pref_upper]
                ordered_headers.append(original_header)
                used_headers.add(pref_upper)
        
        remaining = [h for h_upper, h in all_headers_dict_upper.items() if h_upper not in used_headers]
        ordered_headers.extend(sorted(remaining, key=lambda x: x.upper()))
        
        # Extract data rows
        items = []
        for row_idx in range(header_row + 1, ws.max_row + 1):
            row_data = OrderedDict()
            has_data = False
            
            for header_name in ordered_headers:
                col_idx = None
                for col, h in headers.items():
                    if h == header_name:
                        col_idx = col
                        break
                
                if col_idx:
                    cell_value = ws.cell(row_idx, col_idx).value
                    row_data[header_name] = cell_value
                    if cell_value is not None and str(cell_value).strip() != '':
                        has_data = True
            
            if has_data:
                items.append(row_data)
        
        # Find summary data (NET TOTAL, NO OF UNITS)
        summary = {}
        for row_idx in range(max(1, ws.max_row - 50), ws.max_row + 1):
            row = ws[row_idx]
            if len(row) > 2 and row[2].value:
                item_str = str(row[2].value).upper().strip()
                if 'NET TOTAL' in item_str and 'net_total' not in summary:
                    for check_col in [5, 4, 6, 3]:
                        if len(row) > check_col:
                            value = row[check_col].value
                            if value is not None:
                                try:
                                    if isinstance(value, str):
                                        num_match = re.search(r'[\d,]+\.?\d*', value.replace(',', ''))
                                        if num_match:
                                            value = float(num_match.group())
                                        else:
                                            continue
                                    else:
                                        value = float(value)
                                    summary['net_total'] = value
                                    break
                                except (ValueError, TypeError):
                                    continue
                elif ('NO OF UNITS' in item_str or 'NO OF ITEMS' in item_str) and 'no_of_units' not in summary:
                    for check_col in [5, 4, 6, 3]:
                        if len(row) > check_col:
                            value = row[check_col].value
                            if value is not None:
                                try:
                                    if isinstance(value, str):
                                        num_match = re.search(r'[\d,]+\.?\d*', value.replace(',', ''))
                                        if num_match:
                                            value = float(num_match.group())
                                        else:
                                            continue
                                    else:
                                        value = float(value)
                                    summary['no_of_units'] = value
                                    break
                                except (ValueError, TypeError):
                                    continue
        
        # Convert OrderedDict items to regular dicts
        items_dicts = []
        for item in items:
            if isinstance(item, OrderedDict):
                items_dicts.append(dict(item))
            else:
                items_dicts.append(item)
        
        return {
            'name': board_name,
            'metadata': board_metadata,
            'summary': summary,
            'items': items_dicts
        }
        
    except Exception as e:
        return {'error': str(e)}

def main():
    print("Loading Excel file...")
    try:
        wb = openpyxl.load_workbook('e2.xlsx', data_only=True)
    except FileNotFoundError:
        print("Error: e2.xlsx not found in current directory")
        return
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        return
    
    # Load all boards from TOTALLIST
    print("Loading board list from TOTALLIST sheet...")
    boards = []
    if 'TOTALLIST' in wb.sheetnames:
        ws_totallist = wb['TOTALLIST']
        for row_idx in range(2, ws_totallist.max_row + 1):
            board_name = ws_totallist.cell(row_idx, 5).value  # Column E (Itemdrop)
            if board_name and str(board_name).strip():
                boards.append(str(board_name).strip())
    else:
        print("Error: TOTALLIST sheet not found")
        return
    
    print(f"Found {len(boards)} boards")
    
    # Create board_details directory
    os.makedirs('board_details', exist_ok=True)
    
    # Extract details for each board
    success_count = 0
    error_count = 0
    
    for i, board_name in enumerate(boards, 1):
        print(f"[{i}/{len(boards)}] Processing {board_name}...", end=' ', flush=True)
        
        details = extract_board_details(board_name, wb)
        
        if 'error' in details:
            print(f"ERROR: {details['error']}")
            error_count += 1
        else:
            # Save as JSON file (sanitize filename)
            safe_name = board_name.replace('/', '_').replace('\\', '_').replace(':', '_')
            json_path = os.path.join('board_details', f'{safe_name}.json')
            
            try:
                with open(json_path, 'w', encoding='utf-8') as f:
                    json.dump(details, f, indent=2, default=str, ensure_ascii=False)
                print("OK")
                success_count += 1
            except Exception as e:
                print(f"ERROR saving: {e}")
                error_count += 1
    
    print(f"\nCompleted: {success_count} successful, {error_count} errors")
    print(f"JSON files saved in: board_details/")

if __name__ == '__main__':
    main()

